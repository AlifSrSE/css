from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ResponseMixin:
    """Mixin for standardized API responses"""
    
    def success_response(self, data=None, message="Success", status_code=status.HTTP_200_OK):
        """Return standardized success response"""
        response_data = {
            'success': True,
            'message': message,
            'timestamp': datetime.utcnow().isoformat(),
        }
        
        if data is not None:
            response_data['data'] = data
            
        return Response(response_data, status=status_code)
    
    def error_response(self, message="Error", errors=None, status_code=status.HTTP_400_BAD_REQUEST):
        """Return standardized error response"""
        response_data = {
            'success': False,
            'message': message,
            'timestamp': datetime.utcnow().isoformat(),
        }
        
        if errors:
            response_data['errors'] = errors
            
        return Response(response_data, status=status_code)
    
    def validation_error_response(self, serializer):
        """Return validation error response"""
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors,
            status_code=status.HTTP_400_BAD_REQUEST
        )

class AuditMixin:
    """Mixin for audit trail functionality"""
    
    def log_user_activity(self, user, action, resource=None, details=None, request=None):
        """Log user activity"""
        try:
            from apps.authentication.models import UserActivity
            
            activity_data = {
                'user': user,
                'action': action,
                'timestamp': datetime.utcnow(),
            }
            
            if resource:
                activity_data['resource'] = resource
                
            if details:
                activity_data['details'] = details
                
            if request:
                activity_data['ip_address'] = self.get_client_ip(request)
                activity_data['user_agent'] = request.META.get('HTTP_USER_AGENT', '')
            
            UserActivity(**activity_data).save()
            
        except Exception as e:
            logger.error(f"Failed to log user activity: {str(e)}")
    
    def get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class CacheMixin:
    """Mixin for caching functionality"""
    
    def get_cache_key(self, *args, **kwargs):
        """Generate cache key"""
        key_parts = [str(arg) for arg in args]
        for k, v in kwargs.items():
            key_parts.append(f"{k}:{v}")
        return ":".join(key_parts)
    
    def cache_get(self, key, default=None):
        """Get value from cache"""
        try:
            from django.core.cache import cache
            return cache.get(key, default)
        except Exception as e:
            logger.error(f"Cache get error: {str(e)}")
            return default
    
    def cache_set(self, key, value, timeout=300):
        """Set value in cache"""
        try:
            from django.core.cache import cache
            cache.set(key, value, timeout)
            return True
        except Exception as e:
            logger.error(f"Cache set error: {str(e)}")
            return False
    
    def cache_delete(self, key):
        """Delete value from cache"""
        try:
            from django.core.cache import cache
            cache.delete(key)
            return True
        except Exception as e:
            logger.error(f"Cache delete error: {str(e)}")
            return False

class ValidationMixin:
    """Mixin for common validation functions"""
    
    def validate_required_fields(self, data, required_fields):
        """Validate that all required fields are present"""
        missing_fields = []
        
        for field in required_fields:
            if field not in data or data[field] is None or data[field] == '':
                missing_fields.append(field)
        
        if missing_fields:
            return {
                'is_valid': False,
                'errors': {
                    'missing_fields': missing_fields,
                    'message': f"Required fields missing: {', '.join(missing_fields)}"
                }
            }
        
        return {'is_valid': True}
    
    def validate_numeric_range(self, value, min_val=None, max_val=None, field_name="value"):
        """Validate numeric value within range"""
        try:
            numeric_value = float(value)
            
            if min_val is not None and numeric_value < min_val:
                return {
                    'is_valid': False,
                    'error': f"{field_name} must be at least {min_val}"
                }
            
            if max_val is not None and numeric_value > max_val:
                return {
                    'is_valid': False,
                    'error': f"{field_name} must not exceed {max_val}"
                }
            
            return {'is_valid': True, 'value': numeric_value}
            
        except (ValueError, TypeError):
            return {
                'is_valid': False,
                'error': f"{field_name} must be a valid number"
            }
    
    def validate_date_format(self, date_string, format_string='%Y-%m-%d'):
        """Validate date format"""
        try:
            parsed_date = datetime.strptime(date_string, format_string)
            return {'is_valid': True, 'date': parsed_date}
        except ValueError:
            return {
                'is_valid': False,
                'error': f"Date must be in format {format_string}"
            }

class ExportMixin:
    """Mixin for data export functionality"""
    
    def export_to_csv(self, data, filename, headers=None):
        """Export data to CSV format"""
        import csv
        from io import StringIO
        from django.http import HttpResponse
        
        output = StringIO()
        writer = csv.writer(output)
        
        if headers:
            writer.writerow(headers)
        
        for row in data:
            if isinstance(row, dict):
                writer.writerow(row.values())
            else:
                writer.writerow(row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def export_to_excel(self, data, filename, sheet_name='Sheet1', headers=None):
        """Export data to Excel format"""
        try:
            import openpyxl
            from io import BytesIO
            from django.http import HttpResponse
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            row_num = 1
            
            if headers:
                for col_num, header in enumerate(headers, 1):
                    worksheet.cell(row=row_num, column=col_num, value=header)
                row_num += 1
            
            for row_data in data:
                if isinstance(row_data, dict):
                    values = row_data.values()
                else:
                    values = row_data
                    
                for col_num, value in enumerate(values, 1):
                    worksheet.cell(row=row_num, column=col_num, value=value)
                row_num += 1
            
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            logger.error("openpyxl not installed - cannot export to Excel")
            raise Exception("Excel export not available")
